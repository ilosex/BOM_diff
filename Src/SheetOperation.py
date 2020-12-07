from openpyxl import load_workbook
import xlsxwriter
from Src.component import ParsedComponent, SavedComponent


class Sheet:
    def __init__(self, path):
        self.path = path
        self.header = ''
        self.header_names_en = ['Comment', 'Description', 'Designator', 'Quantity', 'Manufacturer']
        self.header_names_ru = ['Наименование', 'Описание', 'Обозначение', 'Количество', 'Производитель']

    def read_xlsx(self):
        wb = load_workbook(str(self.path))
        return wb.active

    def header_dict_determine(self):
        if 'Comment' in self.header:
            return self.header_names_en
        elif 'Наименование' in self.header:
            return self.header_names_ru
        else:
            print('Язык не определен!')

    def component_compose(self, row):
        header_dict = self.header_dict_determine()
        parameters = list()
        for name in header_dict:
           parameters.append(row[self.header.index(name)]) if name in self.header else parameters.append(None)
        element = ParsedComponent(parameters)
        return element

    def worksheet_parser(self):
        cells = list()
        ws = self.read_xlsx()
        got_header = False
        for row in ws.values:
            if got_header:
                new_element = self.component_compose(row)
                cells.append(new_element)
            elif 'Comment' in row or 'Наименование' in row:
                self.header = row
                got_header = True
        cells.sort()
        return cells


class SheetFormer:
    def __init__(self, path):
        self.path = path
        self.header_names_en = ['Component', 'Quantity', 'Commentary']

    def element_translation(self, parsed_element, comment):
        params = [parsed_element.name, parsed_element.quantity, comment]
        return SavedComponent(params)

    def save_sheet(self, saved_array):
        book_to_save = xlsxwriter.Workbook(str(self.path))
        sheet_to_save = book_to_save.add_worksheet('Diff')
        row_len = len(saved_array)
        for i in range(row_len):
            columns_len = len(saved_array[i])
            for j in range(columns_len):
                sheet_to_save.write(i, j, saved_array[i][j])
        book_to_save.close()
